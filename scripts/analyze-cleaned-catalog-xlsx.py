#!/usr/bin/env python3
"""Compare an edited catalog workbook with a production product snapshot.

This script is intentionally read-only. It creates a machine-readable report and
a candidate patch file; it never connects to or mutates the production database.
"""

from __future__ import annotations

import argparse
import collections
import hashlib
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TEXT_FIELDS = {
    "name": "name",
    "short_description": "shortDescription",
    "full_description": "fullDescription",
    "search_keywords": "searchKeywords",
    "buyer_hint": "buyerHint",
    "brand": "brand",
    "google_product_category": "googleProductCategory",
    "fb_product_category": "fbProductCategory",
    "unit_name": "unitName",
    "package_type": "packageType",
    "package_unit": "packageUnit",
}

NUMBER_FIELDS = {
    "price_with_vat": "priceWithVat",
    "units_per_package": "unitsPerPackage",
    "min_order_packages": "minOrderPackages",
    "sort_order": "sortOrder",
}

BOOL_FIELDS = {
    "is_featured": "isFeatured",
    "is_active": "isActive",
}

IDENTITY_FIELDS = {
    "slug": "slug",
    "meta_catalog_id": "metaCatalogId",
    "image_url": "imageUrl",
}

DIAGNOSTIC_COLUMNS = {
    "original_name",
    "image_sha256",
    "duplicate_status",
    "review_status",
    "review_reason",
}

GENERIC_NAMES = {
    "товар",
    "перчатки",
    "пакет",
    "пакеты",
    "мешок",
    "мешки",
    "коврик",
    "коврики",
}

TECHNICAL_TEXT_RE = re.compile(
    r"(?:Источник:\s*output|WhatsApp\s+product_id|image_sha256|Требует\s+ручной\s+проверки)",
    re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("snapshot", type=Path)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--patch", type=Path, required=True)
    return parser.parse_args()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def clean_multiline_text(value: Any) -> str | None:
    if value is None:
        return None
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in str(value).splitlines()]
    text = "\n".join(line for line in lines if line)
    return text or None


def parse_number(value: Any, *, integer: bool = False) -> int | float | None:
    if value is None or clean_text(value) is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"boolean is not a number: {value!r}")
    number = float(str(value).replace(" ", "").replace(",", "."))
    if not math.isfinite(number):
        raise ValueError(f"non-finite number: {value!r}")
    if integer:
        if not number.is_integer():
            raise ValueError(f"expected integer: {value!r}")
        return int(number)
    return int(number) if number.is_integer() else number


def parse_bool(value: Any) -> bool | None:
    if value is None or clean_text(value) is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "да", "истина"}:
        return True
    if text in {"0", "false", "no", "нет", "ложь"}:
        return False
    raise ValueError(f"invalid boolean: {value!r}")


def parse_characteristics(value: Any) -> dict[str, str] | None:
    """Parse the workbook's `Key: value; Key: value` representation safely."""
    text = clean_multiline_text(value)
    if not text:
        return None
    if text.startswith("{"):
        parsed = json.loads(text)
        if not isinstance(parsed, dict):
            raise ValueError("characteristics JSON must be an object")
        result = {
            clean_text(key): clean_text(item)
            for key, item in parsed.items()
            if clean_text(key) and clean_text(item)
        }
        return dict(result) or None

    segments: list[str] = []
    for line in text.splitlines():
        segments.extend(piece.strip() for piece in line.split(";") if piece.strip())
    result: dict[str, str] = {}
    for segment in segments:
        if ":" not in segment:
            raise ValueError(f"characteristic without colon: {segment!r}")
        key, item = segment.split(":", 1)
        key_text = clean_text(key)
        item_text = clean_text(item)
        if not key_text or not item_text:
            raise ValueError(f"empty characteristic key/value: {segment!r}")
        if key_text in result and result[key_text] != item_text:
            raise ValueError(f"duplicate characteristic key: {key_text!r}")
        result[key_text] = item_text
    return result or None


def comparable(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, dict):
        return {str(k): comparable(v) for k, v in sorted(value.items())}
    return value


def read_sheet(ws: Any) -> list[dict[str, Any]]:
    headers = [clean_text(cell.value) for cell in ws[1]]
    if not all(headers):
        raise ValueError(f"sheet {ws.title!r} has an empty header")
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def current_value(item: dict[str, Any], field: str) -> Any:
    value = item.get(field)
    if field == "unitName" and value is None:
        value = item.get("unit")
    if field == "unitsPerPackage" and value is None:
        value = item.get("packageQuantity")
    if field == "imageUrl" and value is None:
        value = item.get("photo")
    if field == "characteristics" and isinstance(value, str):
        try:
            decoded = json.loads(value)
            if isinstance(decoded, dict):
                value = decoded
        except json.JSONDecodeError:
            pass
    return comparable(value)


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    snapshot_path = args.snapshot.resolve()
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)

    required_sheets = {"Сводка", "Товары", "Ручная проверка", "Удалённые дубли", "Правила"}
    missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
    if missing_sheets:
        raise ValueError(f"missing sheets: {missing_sheets}")

    product_rows = read_sheet(workbook["Товары"])
    review_rows = read_sheet(workbook["Ручная проверка"])
    duplicate_rows = read_sheet(workbook["Удалённые дубли"])

    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    production_items = snapshot.get("items") or []
    production_by_external_id = {
        clean_text(item.get("externalId")): item
        for item in production_items
        if clean_text(item.get("externalId"))
    }

    ids = [clean_text(row.get("external_id")) for row in product_rows]
    id_counts = collections.Counter(ids)
    duplicate_input_ids = sorted(key for key, count in id_counts.items() if key and count > 1)
    missing_input_ids = [index + 2 for index, value in enumerate(ids) if not value]
    if duplicate_input_ids or missing_input_ids:
        raise ValueError(
            f"invalid external_id values: duplicates={duplicate_input_ids}, missing_rows={missing_input_ids}"
        )

    category_ids: dict[str, set[str]] = collections.defaultdict(set)
    subcategory_ids: dict[tuple[str, str], set[str]] = collections.defaultdict(set)
    for item in production_items:
        category = item.get("category") or {}
        subcategory = item.get("subcategory") or {}
        category_name = clean_text(category.get("name"))
        subcategory_name = clean_text(subcategory.get("name"))
        if category_name and category.get("id"):
            category_ids[category_name].add(category["id"])
        if category_name and subcategory_name and subcategory.get("id"):
            subcategory_ids[(category_name, subcategory_name)].add(subcategory["id"])

    field_change_counts: collections.Counter[str] = collections.Counter()
    validation_counts: collections.Counter[str] = collections.Counter()
    advisory_counts: collections.Counter[str] = collections.Counter()
    candidate_items: list[dict[str, Any]] = []
    ambiguous_items: list[dict[str, Any]] = []
    unchanged_ids: list[str] = []
    new_ids: list[str] = []

    for row in product_rows:
        external_id = clean_text(row.get("external_id"))
        assert external_id
        existing = production_by_external_id.get(external_id)
        if not existing:
            new_ids.append(external_id)
            ambiguous_items.append({
                "externalId": external_id,
                "reasons": ["not_found_in_production"],
            })
            continue

        reasons: list[str] = []
        advisories: list[str] = []
        target: dict[str, Any] = {}
        identity_differences: dict[str, dict[str, Any]] = {}

        for workbook_field, database_field in TEXT_FIELDS.items():
            raw = row.get(workbook_field)
            value = clean_multiline_text(raw) if database_field in {"shortDescription", "fullDescription"} else clean_text(raw)
            # Blank optional cells must never erase surviving production data implicitly.
            if value is not None:
                target[database_field] = value

        for workbook_field, database_field in NUMBER_FIELDS.items():
            try:
                value = parse_number(
                    row.get(workbook_field),
                    integer=database_field in {"unitsPerPackage", "minOrderPackages", "sortOrder"},
                )
            except ValueError as exc:
                reasons.append(f"{workbook_field}: {exc}")
                continue
            if value is not None:
                target[database_field] = value

        for workbook_field, database_field in BOOL_FIELDS.items():
            try:
                value = parse_bool(row.get(workbook_field))
            except ValueError as exc:
                reasons.append(f"{workbook_field}: {exc}")
                continue
            if value is not None:
                target[database_field] = value

        try:
            characteristics = parse_characteristics(row.get("characteristics"))
            if characteristics is not None:
                target["characteristics"] = characteristics
        except (ValueError, json.JSONDecodeError) as exc:
            reasons.append(f"characteristics: {exc}")

        category_name = clean_text(row.get("category"))
        subcategory_name = clean_text(row.get("subcategory"))
        classification: dict[str, Any] = {
            "categoryName": category_name,
            "subcategoryName": subcategory_name,
            "categoryId": None,
            "subcategoryId": None,
            "createSubcategory": False,
        }
        if not category_name:
            reasons.append("category_missing")
        elif len(category_ids.get(category_name, set())) != 1:
            reasons.append(f"category_not_uniquely_resolved: {category_name}")
        else:
            category_id = next(iter(category_ids[category_name]))
            target["categoryId"] = category_id
            classification["categoryId"] = category_id

        if subcategory_name:
            subcategory_key = (category_name or "", subcategory_name)
            resolved_subcategories = subcategory_ids.get(subcategory_key, set())
            if len(resolved_subcategories) > 1:
                reasons.append(f"subcategory_not_uniquely_resolved: {category_name} / {subcategory_name}")
            elif len(resolved_subcategories) == 0:
                classification["createSubcategory"] = True
                advisories.append(f"create_subcategory: {category_name} / {subcategory_name}")
            else:
                subcategory_id = next(iter(resolved_subcategories))
                target["subcategoryId"] = subcategory_id
                classification["subcategoryId"] = subcategory_id

        for workbook_field, database_field in IDENTITY_FIELDS.items():
            incoming = clean_text(row.get(workbook_field))
            current = clean_text(current_value(existing, database_field))
            if incoming and incoming != current:
                identity_differences[database_field] = {"current": current, "incoming": incoming}

        name = target.get("name")
        if not name:
            reasons.append("name_missing")
        elif name.casefold() in GENERIC_NAMES or len(name) < 5:
            advisories.append(f"name_too_generic: {name}")

        short_description = target.get("shortDescription")
        full_description = target.get("fullDescription")
        if short_description:
            # Keep the legacy description field buyer-safe too: the search and
            # quality checker still inspect it.
            target["description"] = short_description
            if not full_description or TECHNICAL_TEXT_RE.search(full_description):
                target["fullDescription"] = short_description
                advisories.append("technical_full_description_replaced")

        incoming_price = target.get("priceWithVat")
        current_price = current_value(existing, "priceWithVat")
        if incoming_price != current_price:
            reasons.append(f"price_changed: {current_price!r} -> {incoming_price!r}")

        if target.get("isActive") != current_value(existing, "isActive"):
            reasons.append(
                f"active_state_changed: {current_value(existing, 'isActive')!r} -> {target.get('isActive')!r}"
            )

        changes: dict[str, dict[str, Any]] = {}
        for field, incoming in target.items():
            current = current_value(existing, field)
            incoming = comparable(incoming)
            if current != incoming:
                changes[field] = {"from": current, "to": incoming}
                field_change_counts[field] += 1

        if identity_differences:
            validation_counts["identity_differences"] += 1
        for reason in reasons:
            validation_counts[reason.split(":", 1)[0]] += 1
        for advisory in advisories:
            advisory_counts[advisory.split(":", 1)[0]] += 1

        current_category_name = clean_text((existing.get("category") or {}).get("name"))
        current_subcategory_name = clean_text((existing.get("subcategory") or {}).get("name"))
        classification_changed = (
            current_category_name != category_name
            or current_subcategory_name != subcategory_name
        )

        if reasons:
            ambiguous_items.append({
                "externalId": external_id,
                "name": target.get("name"),
                "reasons": reasons,
                "identityDifferences": identity_differences,
                "changes": changes,
                "advisories": advisories,
                "classification": classification,
            })
        elif changes or classification_changed:
            candidate_items.append({
                "id": existing["id"],
                "externalId": external_id,
                "expectedUpdatedAt": existing.get("updatedAt"),
                "changes": {field: details["to"] for field, details in changes.items()},
                "changeSummary": changes,
                "classification": classification,
                "classificationChanged": classification_changed,
                "advisories": advisories,
            })
        else:
            unchanged_ids.append(external_id)

    workbook_ids = set(ids)
    omitted_production_ids = sorted(set(production_by_external_id) - workbook_ids)
    declared_duplicate_ids = sorted(
        clean_text(row.get("external_id"))
        for row in duplicate_rows
        if clean_text(row.get("external_id"))
    )
    undeclared_omissions = sorted(set(omitted_production_ids) - set(declared_duplicate_ids))
    declared_but_present = sorted(set(declared_duplicate_ids) & workbook_ids)

    review_status_counts = collections.Counter(clean_text(row.get("review_status")) for row in product_rows)
    manual_review_status_counts = collections.Counter(clean_text(row.get("review_status")) for row in review_rows)
    review_directives: list[dict[str, Any]] = []
    for row in review_rows:
        external_id = clean_text(row.get("external_id"))
        if not external_id or external_id not in production_by_external_id:
            continue
        reason = clean_multiline_text(row.get("review_reason"))
        source_status = clean_text(row.get("review_status"))
        note_parts = ["Импортировано из проверенного Excel-файла."]
        if source_status:
            note_parts.append(f"Статус: {source_status}.")
        if reason:
            note_parts.append(reason)
        review_directives.append({
            "externalId": external_id,
            "status": "PENDING",
            "note": " ".join(note_parts)[:2000],
        })

    patch = {
        "schemaVersion": 1,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": {
            "workbook": str(workbook_path),
            "workbookSha256": sha256(workbook_path),
            "snapshot": str(snapshot_path),
            "snapshotSha256": sha256(snapshot_path),
            "snapshotGeneratedAt": snapshot.get("generatedAt"),
        },
        "policy": {
            "deleteAbsentRows": False,
            "preserveBlankOptionalFields": True,
            "blockPriceChanges": True,
            "blockActiveStateChanges": True,
            "requireExistingCategory": True,
            "createMissingSubcategoriesWithinExistingCategories": True,
            "replaceTechnicalFullDescriptionWithCleanShortDescription": True,
        },
        "items": candidate_items,
        "reviews": review_directives,
    }

    report = {
        "generatedAt": patch["generatedAt"],
        "source": patch["source"],
        "counts": {
            "production": len(production_items),
            "workbookProducts": len(product_rows),
            "workbookReviewRows": len(review_rows),
            "workbookDeclaredDuplicates": len(duplicate_rows),
            "safeCandidateUpdates": len(candidate_items),
            "unchanged": len(unchanged_ids),
            "ambiguous": len(ambiguous_items),
            "new": len(new_ids),
            "omittedFromWorkbook": len(omitted_production_ids),
            "pendingReviewDirectives": len(review_directives),
        },
        "fieldChangeCounts": dict(field_change_counts.most_common()),
        "validationCounts": dict(validation_counts.most_common()),
        "advisoryCounts": dict(advisory_counts.most_common()),
        "reviewStatusCounts": {
            "products": {str(key): value for key, value in review_status_counts.most_common()},
            "manualSheet": {str(key): value for key, value in manual_review_status_counts.most_common()},
        },
        "omissions": {
            "all": omitted_production_ids,
            "declaredDuplicates": declared_duplicate_ids,
            "undeclared": undeclared_omissions,
            "declaredButStillPresent": declared_but_present,
            "action": "preserve_in_production",
        },
        "newIds": new_ids,
        "unchangedIds": unchanged_ids,
        "ambiguousItems": ambiguous_items,
        "diagnosticColumnsIgnored": sorted(DIAGNOSTIC_COLUMNS),
    }

    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.patch.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    args.patch.write_text(json.dumps(patch, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({
        "counts": report["counts"],
        "fieldChangeCounts": report["fieldChangeCounts"],
        "validationCounts": report["validationCounts"],
        "advisoryCounts": report["advisoryCounts"],
        "reviewStatusCounts": report["reviewStatusCounts"],
        "omissions": report["omissions"],
        "report": str(args.report.resolve()),
        "patch": str(args.patch.resolve()),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
